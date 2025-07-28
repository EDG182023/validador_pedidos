from typing import List
from io import BytesIO
from openpyxl import load_workbook
from ..models.dtos import PedidoSaadisDto

class ExcelService:
    """Simple parser for Excel files."""

    def leer_pedidos(self, data: bytes) -> List[PedidoSaadisDto]:
        pedidos: List[PedidoSaadisDto] = []
        wb = load_workbook(filename=BytesIO(data))
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            pedido = PedidoSaadisDto(
                rhd_nrocli=row[0],
                cbt_codcbt=row[1],
                cbt_cenemi=row[2],
                cbt_nrocbt=row[3],
                cbt_letcbt=row[4],
                cbt_feccbt=row[5],
                cbt_detall=row[6] or "",
                detalle=[],
            )
            pedidos.append(pedido)
        return pedidos
